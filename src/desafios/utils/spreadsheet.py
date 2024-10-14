import pandas as pd
from openpyxl import load_workbook
from IPython.display import display

import os, sys

path_services = os.path.abspath(os.path.join(
    os.path.dirname(__file__), '..', 'services'
))
sys.path.append(path_services)

from services import Product as product

def read_excel(path_file, name_spreadsheet):
    df = pd.read_excel(path_file, sheet_name=name_spreadsheet)

    return df

def show_data_excel(df):
    display(df)

def update_excel(df):
    # Load the existing workbook
    # workbook = openpyxl.load_workbook(r'C:\Users\matutino\Documents\projects\lg-poo\bot_poo_excel\spreadsheet\product_forms.xlsx')
    workbook = load_workbook(df)

    # Select the sheet you want to work with
    sheet = workbook['Planilha1']

    # Overwrite a cell value
    sheet['A2'] = "Notebook Acer"
    sheet['B2'] = 3500
    sheet['C2'] = 10

    # Save the workbook
    workbook.save(r'C:\Users\matutino\Documents\projects\lg-poo\bot_poo_excel\spreadsheet\product_forms.xlsx')